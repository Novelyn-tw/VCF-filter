#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VCF Variant Analysis Script
Analyzes VCF file to find PASS variants and their associated genes and diseases
Uses Ensembl REST API and other databases programmatically
"""

import requests
import json
import time
import csv
from typing import Dict, List, Optional
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel output will not be available.")
    print("Install with: pip install openpyxl")

class VariantAnalyzer:
    def __init__(self):
        self.ensembl_server = "https://rest.ensembl.org"
        self.clinvar_base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
        
    def parse_vcf_line(self, line: str) -> Dict:
        """Parse a single VCF data line"""
        fields = line.strip().split('\t')
        if len(fields) < 8:
            return None
            
        info_field = fields[7]
        info_dict = {}
        for item in info_field.split(';'):
            if '=' in item:
                key, value = item.split('=', 1)
                info_dict[key] = value
                
        return {
            'chrom': fields[0],
            'pos': int(fields[1]),
            'id': fields[2] if fields[2] != '.' else None,
            'ref': fields[3],
            'alt': fields[4],
            'qual': fields[5],
            'filter': fields[6],
            'info': info_dict
        }
    
    def get_gene_from_position(self, chrom: str, pos: int) -> Optional[Dict]:
        """Get gene information from genomic position using Ensembl API"""
        # Remove 'chr' prefix if present for Ensembl API
        chrom_clean = chrom.replace('chr', '')
        url = f"{self.ensembl_server}/overlap/region/human/{chrom_clean}:{pos}-{pos}?feature=gene;content-type=application/json"
        
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                genes = response.json()
                if genes:
                    gene = genes[0]
                    return {
                        'gene_id': gene.get('id'),
                        'gene_name': gene.get('external_name'),
                        'description': gene.get('description', ''),
                        'biotype': gene.get('biotype')
                    }
        except Exception as e:
            print(f"Error fetching gene for {chrom}:{pos} - {e}")
            
        return None
    
    def get_variant_info_from_ensembl(self, chrom: str, pos: int, ref: str, alt: str) -> Optional[Dict]:
        """Get variant information from Ensembl using position and alleles"""
        chrom_clean = chrom.replace('chr', '')
        
        # Try to get variant by position first
        url = f"{self.ensembl_server}/variation/human/{chrom_clean}:{pos}-{pos}:1?content-type=application/json"
        
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data:
                    for variant in data:
                        mappings = variant.get('mappings', [])
                        for mapping in mappings:
                            if (mapping.get('start') == pos and 
                                mapping.get('allele_string', '').split('/')[0] == ref):
                                
                                return {
                                    'clinical_significance': variant.get('clinical_significance', []),
                                    'phenotypes': variant.get('phenotypes', []),
                                    'minor_allele_freq': variant.get('minor_allele_freq'),
                                    'synonyms': variant.get('synonyms', [])
                                }
        except Exception as e:
            print(f"Error fetching variant info for {chrom}:{pos} - {e}")
            
        return None
    
    def get_variant_info_from_dbsnp(self, rs_id: str) -> Optional[Dict]:
        """Get variant information from dbSNP using rs ID"""
        if not rs_id or not rs_id.startswith('rs'):
            return None
            
        url = f"{self.ensembl_server}/variation/human/{rs_id}?content-type=application/json"
        
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                return {
                    'clinical_significance': data.get('clinical_significance', []),
                    'phenotypes': data.get('phenotypes', []),
                    'minor_allele_freq': data.get('minor_allele_freq'),
                    'synonyms': data.get('synonyms', [])
                }
        except Exception as e:
            print(f"Error fetching variant info for {rs_id} - {e}")
            
        return None
    
    def search_clinvar_by_position(self, chrom: str, pos: int) -> List[str]:
        """Search ClinVar for disease associations using genomic position"""
        chrom_clean = chrom.replace('chr', '')
        search_term = f"{chrom_clean}[chr] AND {pos}[chrpos37]"
        
        search_url = f"{self.clinvar_base}/esearch.fcgi"
        search_params = {
            'db': 'clinvar',
            'term': search_term,
            'retmode': 'json',
            'retmax': 10
        }
        
        try:
            response = requests.get(search_url, params=search_params, timeout=10)
            if response.status_code == 200:
                search_data = response.json()
                id_list = search_data.get('esearchresult', {}).get('idlist', [])
                
                if id_list:
                    return self._fetch_clinvar_diseases(id_list)
                        
        except Exception as e:
            print(f"Error searching ClinVar for {chrom}:{pos} - {e}")
            
        return []
    
    def search_clinvar_by_rsid(self, rs_id: str) -> List[str]:
        """Search ClinVar for disease associations using rs ID"""
        if not rs_id or not rs_id.startswith('rs'):
            return []
            
        search_url = f"{self.clinvar_base}/esearch.fcgi"
        search_params = {
            'db': 'clinvar',
            'term': rs_id,
            'retmode': 'json',
            'retmax': 10
        }
        
        try:
            response = requests.get(search_url, params=search_params, timeout=10)
            if response.status_code == 200:
                search_data = response.json()
                id_list = search_data.get('esearchresult', {}).get('idlist', [])
                
                if id_list:
                    return self._fetch_clinvar_diseases(id_list)
                        
        except Exception as e:
            print(f"Error searching ClinVar for {rs_id} - {e}")
            
        return []
    
    def _fetch_clinvar_diseases(self, id_list: List[str]) -> List[str]:
        """Fetch disease information from ClinVar IDs"""
        fetch_url = f"{self.clinvar_base}/esummary.fcgi"
        fetch_params = {
            'db': 'clinvar',
            'id': ','.join(id_list[:5]),
            'retmode': 'json'
        }
        
        try:
            fetch_response = requests.get(fetch_url, params=fetch_params, timeout=10)
            if fetch_response.status_code == 200:
                fetch_data = fetch_response.json()
                diseases = []
                
                for uid in id_list[:5]:
                    if uid in fetch_data.get('result', {}):
                        title = fetch_data['result'][uid].get('title', '')
                        if title:
                            diseases.append(title)
                
                return diseases
        except Exception as e:
            print(f"Error fetching ClinVar diseases - {e}")
            
        return []
    
    def extract_allele_frequency(self, info_dict: Dict) -> str:
        """Extract allele frequency from INFO field with multiple possible keys"""
        # Common AF keys in VCF files
        af_keys = ['AF', 'MAF', 'CAF', 'GMAF', 'ExAC_AF', 'gnomAD_AF', '1000Gp3_AF']
        
        for key in af_keys:
            if key in info_dict:
                af_value = info_dict[key]
                # Handle comma-separated values (take first one)
                if ',' in af_value:
                    af_value = af_value.split(',')[0]
                try:
                    # Convert to float and format nicely
                    af_float = float(af_value)
                    return f"{af_float:.6f}"
                except ValueError:
                    continue
        
        return "N/A"
    
    def analyze_vcf_file(self, vcf_file_path: str) -> List[Dict]:
        """Analyze VCF file and return results as list of dictionaries"""
        results = []
        
        try:
            with open(vcf_file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.startswith('#'):
                        continue
                        
                    variant = self.parse_vcf_line(line)
                    if not variant:
                        continue
                    
                    if variant['filter'] != 'PASS':
                        continue
                    
                    print(f"Analyzing variant: {variant['chrom']}:{variant['pos']} {variant['ref']}>{variant['alt']}")
                    
                    # Get gene information
                    gene_info = self.get_gene_from_position(variant['chrom'], variant['pos'])
                    
                    # Get variant information from multiple sources
                    variant_info = None
                    if variant['id']:
                        variant_info = self.get_variant_info_from_dbsnp(variant['id'])
                    
                    if not variant_info:
                        variant_info = self.get_variant_info_from_ensembl(
                            variant['chrom'], variant['pos'], variant['ref'], variant['alt']
                        )
                    
                    # Get disease associations
                    diseases = []
                    if variant['id']:
                        diseases = self.search_clinvar_by_rsid(variant['id'])
                    
                    if not diseases:
                        diseases = self.search_clinvar_by_position(variant['chrom'], variant['pos'])
                    
                    # Extract allele frequency
                    allele_freq = self.extract_allele_frequency(variant['info'])
                    
                    # If we have variant info but no AF from VCF, try from variant info
                    if allele_freq == "N/A" and variant_info and variant_info.get('minor_allele_freq'):
                        allele_freq = str(variant_info['minor_allele_freq'])
                    
                    # Extract clinical significance
                    clinical_sig = "Unknown"
                    if variant_info and variant_info.get('clinical_significance'):
                        clinical_sig = ', '.join(variant_info['clinical_significance'])
                    
                    result = {
                        'chromosome': variant['chrom'],
                        'position': variant['pos'],
                        'rs_id': variant['id'] or 'N/A',
                        'ref_allele': variant['ref'],
                        'alt_allele': variant['alt'],
                        'gene_name': gene_info['gene_name'] if gene_info else 'Unknown',
                        'gene_id': gene_info['gene_id'] if gene_info else 'Unknown',
                        'gene_description': gene_info['description'] if gene_info else 'Unknown',
                        'allele_frequency': allele_freq,
                        'clinical_significance': clinical_sig,
                        'associated_diseases': '; '.join(diseases) if diseases else 'No disease associations found'
                    }
                    
                    results.append(result)
                    time.sleep(0.5)  # Rate limiting
                    
        except FileNotFoundError:
            print(f"Error: VCF file '{vcf_file_path}' not found")
            return []
        except Exception as e:
            print(f"Error analyzing VCF file: {e}")
            return []
        
        return results

def save_to_excel(results: List[Dict], filename: str):
    """Save results to a simple Excel file"""
    if not results or not EXCEL_AVAILABLE:
        print("Cannot create Excel file. Results will be saved as CSV instead.")
        save_to_csv(results, filename.replace('.xlsx', '.csv'))
        return
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variant Analysis Results"
    
    # Define headers
    headers = [
        'Chromosome', 'Position', 'RS_ID', 'Reference', 'Alternate', 
        'Gene_Name', 'Gene_ID', 'Gene_Description', 'Allele_Frequency', 
        'Clinical_Significance', 'Associated_Diseases'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data rows
    for row_idx, result in enumerate(results, 2):
        data_row = [
            f"chr{result['chromosome']}",
            result['position'],
            result['rs_id'],
            result['ref_allele'],
            result['alt_allele'],
            result['gene_name'],
            result['gene_id'],
            result['gene_description'],
            result['allele_frequency'],
            result['clinical_significance'],
            result['associated_diseases']
        ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Adjust column widths
    column_widths = [12, 12, 15, 10, 10, 15, 20, 40, 12, 25, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Add summary worksheet
    summary_ws = wb.create_sheet(title="Summary")
    
    # Summary data
    summary_data = [
        ["Analysis Summary", ""],
        ["Total PASS variants analyzed", len(results)],
        ["Variants with gene assignments", len([r for r in results if r['gene_name'] != 'Unknown'])],
        ["Variants with disease associations", len([r for r in results if r['associated_diseases'] != 'No disease associations found'])],
        ["Variants with allele frequency", len([r for r in results if r['allele_frequency'] != 'N/A'])],
        ["Variants with clinical significance", len([r for r in results if r['clinical_significance'] != 'Unknown'])],
        ["", ""],
        ["Gene Distribution", ""],
    ]
    
    # Count genes - fix the None issue
    gene_counts = {}
    for result in results:
        gene = result['gene_name']
        if gene and gene != 'Unknown':
            gene_counts[gene] = gene_counts.get(gene, 0) + 1
    
    # Sort genes properly
    for gene in sorted(gene_counts.keys()):
        summary_data.append([gene, gene_counts[gene]])
    
    # Add summary data to worksheet
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=label)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    # Adjust summary column widths
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 15
    
    # Save workbook
    wb.save(filename)
    print(f"Excel file saved: {filename}")

def save_to_csv(results: List[Dict], filename: str):
    """Save results to CSV file as backup"""
    if not results:
        return
        
    fieldnames = [
        'chromosome', 'position', 'rs_id', 'ref_allele', 'alt_allele',
        'gene_name', 'gene_id', 'gene_description', 'allele_frequency',
        'clinical_significance', 'associated_diseases'
    ]
    
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

def main():
    """Main function to run the analysis"""
    analyzer = VariantAnalyzer()
    
    vcf_file = "clinical_test_filtered.vcf"
    print(f"Analyzing VCF file: {vcf_file}")
    print("=" * 60)
    
    results = analyzer.analyze_vcf_file(vcf_file)
    
    if results:
        print(f"\nFound {len(results)} PASS variants")
        print("Processing results...")
        
        # Create Excel output
        excel_file = "clinical_test_annotation_results.xlsx"
        save_to_excel(results, excel_file)
        
        # Also create CSV as backup
        csv_file = "clinical_test_annotation_results.csv"
        save_to_csv(results, csv_file)
        print(f"CSV backup saved: {csv_file}")
        
        # Display clean summary
        print(f"\n" + "="*60)
        print("ANALYSIS SUMMARY")
        print("="*60)
        print(f"Total PASS variants analyzed: {len(results)}")
        
        gene_count = len([r for r in results if r['gene_name'] != 'Unknown'])
        print(f"Variants with gene assignments: {gene_count}")
        
        af_count = len([r for r in results if r['allele_frequency'] != 'N/A'])
        print(f"Variants with allele frequency: {af_count}")
        
        clinical_count = len([r for r in results if r['clinical_significance'] != 'Unknown'])
        print(f"Variants with clinical significance: {clinical_count}")
        
        disease_count = len([r for r in results if r['associated_diseases'] != 'No disease associations found'])
        print(f"Variants with disease associations: {disease_count}")
        
        # Show gene distribution
        gene_counts = {}
        for result in results:
            gene = result['gene_name']
            if gene and gene != 'Unknown':
                gene_counts[gene] = gene_counts.get(gene, 0) + 1
        
        if gene_counts:
            print(f"\nGene Distribution:")
            for gene in sorted(gene_counts.keys()):
                print(f"  {gene}: {gene_counts[gene]} variant(s)")
        
        # Show variants with clinical significance
        clinical_variants = [r for r in results if r['clinical_significance'] != 'Unknown']
        if clinical_variants:
            print(f"\nClinically Significant Variants:")
            for var in clinical_variants:
                print(f"  {var['rs_id']} ({var['gene_name']}): {var['clinical_significance']}")
        
        print(f"\n" + "="*60)
        print("FILES CREATED:")
        print(f"  ? {excel_file} (Main results with formatting)")
        print(f"  ? {csv_file} (Backup CSV format)")
        print("="*60)
        
    else:
        print("No PASS variants found or error in analysis")

if __name__ == "__main__":
    main()